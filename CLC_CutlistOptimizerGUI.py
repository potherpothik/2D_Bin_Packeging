import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from svgwrite import Drawing
import openpyxl
import math
import os
import subprocess
from PIL import Image, ImageTk
from ttkthemes import ThemedTk
import shutil
import sys

def setup_resources():
    # Determine if running as a script or packaged executable
    if getattr(sys, 'frozen', False):
        # Running as a packaged executable (PyInstaller)
        application_path = sys._MEIPASS
    else:
        # Running as a script in a development environment
        application_path = os.path.dirname(os.path.abspath(__file__))

    resources_dir = os.path.join(os.path.abspath("."), "resources")
    if not os.path.exists(resources_dir):
        os.makedirs(resources_dir)
        print("Resources directory created at:", resources_dir)

    resource_files = ["clc.png"]
    for resource_file in resource_files:
        source_path = os.path.join(application_path, resource_file)
        destination_path = os.path.join(resources_dir, resource_file)
        if not os.path.exists(destination_path):
            shutil.copy(source_path, destination_path)
            print(f"Copied: {resource_file} to {resources_dir}")
            
class CutlistOptimizerGUI:
    def __init__(self):
        self.window = ThemedTk(theme="clam")  # Choose your desired theme here
        self.window.title("Cut Layout Creator")
        self.setup_variables()
        self.setup_widgets()  
        self.window.grid_columnconfigure(1, weight=1)
        self.window.resizable(True, False)
        self.window.geometry('700x395')
        self.window.mainloop()
    
    def setup_variables(self):
        # Define the default font using tkinter.font.Font

        self.stock_width = tk.StringVar(value='1220')
        self.stock_length = tk.StringVar(value='2440')
        self.gap = tk.StringVar(value='12.7')
        self.project_id = tk.StringVar(value='Project1')
        self.row_num = 0
        last_output_folder = self.load_last_output_folder_path()
        if last_output_folder:
            self.output_entry = ttk.Entry(self.window, textvariable=tk.StringVar(value=last_output_folder))
        else:
            self.output_entry = ttk.Entry(self.window, textvariable=tk.StringVar())
        self.output_folder_path = tk.StringVar()

    def setup_widgets(self):
        # Insert the image at the desired position
        self.insert_image()

        # Configure the weights of the columns and rows
        self.window.grid_columnconfigure(0, weight=0)  # First column has a fixed width
        self.window.grid_columnconfigure(1, weight=1)  # Middle column expands
        self.window.grid_columnconfigure(2, weight=0)  # Last column has a fixed width

        # Instruction Button``
        self.row_num += 1
        ttk.Button(self.window, text="Instructions", command=self.display_instructions).grid(row=self.row_num, column=0, columnspan=3, pady=10, sticky="ew")

        # Excel Cut List Row
        self.row_num += 1
        ttk.Label(self.window, text="Excel Cut List:").grid(row=self.row_num, column=0, padx=10, pady=10, sticky="e")
        self.file_entry = ttk.Entry(self.window, textvariable=tk.StringVar())
        self.file_entry.grid(row=self.row_num, column=1, padx=10, pady=10, sticky="ew")
        ttk.Button(self.window, text="Select File", command=self.load_cutlist).grid(row=self.row_num, column=2, padx=5, pady=10, sticky="ew")

        # Output Folder Row
        self.row_num += 1
        ttk.Label(self.window, text="Output Folder:").grid(row=self.row_num, column=0, padx=10, pady=10, sticky="e")
        self.output_entry = ttk.Entry(self.window, textvariable=self.output_folder_path)
        self.output_entry.grid(row=self.row_num, column=1, padx=10, pady=10, sticky="ew")
        ttk.Button(self.window, text="Select Folder", command=self.select_output_folder).grid(row=self.row_num, column=2, padx=5, pady=10, sticky="ew")

        # Stock Width Row
        self.row_num += 1
        ttk.Label(self.window, text="Stock Width (mm):").grid(row=self.row_num, column=0, padx=10, pady=10, sticky="e")
        ttk.Entry(self.window, textvariable=self.stock_width).grid(row=self.row_num, column=1, padx=10, pady=10, sticky="ew")

        # Stock Length Row
        self.row_num += 1
        ttk.Label(self.window, text="Stock Length (mm):").grid(row=self.row_num, column=0, padx=10, pady=10, sticky="e")
        ttk.Entry(self.window, textvariable=self.stock_length).grid(row=self.row_num, column=1, padx=10, pady=10, sticky="ew")

        # Gap Between Parts Row
        self.row_num += 1
        ttk.Label(self.window, text="Gap Between Parts (mm):").grid(row=self.row_num, column=0, padx=10, pady=10, sticky="e")
        ttk.Entry(self.window, textvariable=self.gap).grid(row=self.row_num, column=1, padx=10, pady=10, sticky="ew")

        # Project ID Row
        self.row_num += 1
        ttk.Label(self.window, text="Project ID:").grid(row=self.row_num, column=0, padx=10, pady=10, sticky="e")
        ttk.Entry(self.window, textvariable=self.project_id,style="EEntry.TEntry").grid(row=self.row_num, column=1, padx=10, pady=10, sticky="ew")

        # Create & Export Cut List and Open Output Directory Buttons
        self.row_num += 1
        ttk.Button(self.window, text="Create & Export Cut List", command=self.create_export_cutlist).grid(row=self.row_num, column=0, columnspan=2, pady=10, sticky="ew")
        ttk.Button(self.window, text="Open Output Directory", command=self.open_output_directory).grid(row=self.row_num, column=2, pady=10, sticky="ew")

        # Configure the weight of the bottom row to prevent vertical expansion
        self.window.grid_rowconfigure(self.row_num, weight=0)
        last_output_folder = self.load_last_output_folder_path()
        if last_output_folder:
            self.output_folder_path.set(last_output_folder)  # Update the StringVar with the loaded path

    def insert_image(self):
        # Load the icon image
        resources_path= os.path.join(os.path.dirname(__file__), 'resources')
        icon_path = os.path.join(resources_path, 'clc.png')

        self.icon_image = Image.open(icon_path)
        
        # Resize the image to a specific size (e.g., 200x200 pixels)
        self.icon_image_resized = self.icon_image.resize((175, 175), Image.LANCZOS)
        self.icon_photo = ImageTk.PhotoImage(self.icon_image_resized)

        # Create a label to hold the image
        self.icon_label = tk.Label(self.window, image=self.icon_photo)
        self.icon_label.grid(row=4, column=2, rowspan=4, sticky='nsew')

        # Configure row and column weights to allow resizing
        for i in range(5):
            self.window.grid_rowconfigure(i, weight=1)
        for i in range(3):
            self.window.grid_columnconfigure(i, weight=1) 

    def load_cutlist(self):
        self.input_file_path = filedialog.askopenfilename(title="Select Excel Cut List", filetypes=[("Excel Files", "*.xlsx")])
        if self.input_file_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, self.input_file_path)

    def select_output_folder(self):
        selected_folder = filedialog.askdirectory(title="Select Output Folder")
        print(f"Folder selected: {selected_folder}")  # Print the selected folder path
        if selected_folder:
            self.output_folder_path.set(selected_folder)  # Update the StringVar
            print(f"StringVar set to: {self.output_folder_path.get()}")  # Check the updated StringVar value
            self.save_output_folder_path(selected_folder)  # Save the path if needed

    def create_export_cutlist(self):
        if hasattr(self, 'input_file_path') and hasattr(self, 'output_folder_path'):
            try:
                stock_width = float(self.stock_width.get())
                stock_length = float(self.stock_length.get())
                gap = float(self.gap.get())
                cut_list = []  # Initialize the cut list
                wb = openpyxl.load_workbook(self.input_file_path)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    part_label, length, height, count, material = row
                    count = int(count)  # Convert count to integer
                    for _ in range(count):
                        cut_list.append({'Part Label': part_label, 'Length': math.ceil(float(length)), 'Height': math.ceil(float(height)), 'Material': material})
                project_folder = self.determine_project_folder()
                self.create_and_export_cutlists(cut_list, (stock_length, stock_width), gap, project_folder)
                messagebox.showinfo("Success", f"SVG layouts have been created and exported successfully in {project_folder}.")
            except ValueError:
                messagebox.showerror("Error", "Invalid dimensions or gap value. Please enter valid numbers.")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {e}")

    def load_cut_list(self, filepath):
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        return [{'Part Label': row[0], 'Length': math.ceil(float(row[1])), 'Height': math.ceil(float(row[2])), 'Material': row[4]} for row in sheet.iter_rows(min_row=2, values_only=True)]

    def determine_project_folder(self):
        project_id = self.project_id.get().strip()
        output_folder_path_str = self.output_folder_path.get()
        project_folder = os.path.join(output_folder_path_str, project_id) if project_id else os.path.join(output_folder_path_str, "default_project_folder")

       # project_folder = os.path.join(self.output_folder_path, project_id) if project_id else self.get_project_folder(self.output_folder_path)
        os.makedirs(project_folder, exist_ok=True)
        

    def create_and_export_cutlists(self, cut_list, plywood_size, gap, project_folder):
        # Group parts based on material
        materials = {}
        for part in cut_list:
            material = part['Material']
            if material not in materials:
                materials[material] = []
            materials[material].append(part)

        # Create and export cut lists for each material
        for material, parts in materials.items():
            layouts = self.calculate_layout(parts, plywood_size, gap)
            self.create_svg(layouts, project_folder, gap, material)

    def calculate_layout(self, parts, plywood_size, gap):
        parts.sort(key=lambda x: x['Height'], reverse=True)
        sheet_layouts, current_sheet = [], {'parts': [], 'positions': []}
        current_x, current_y = gap, gap
        max_y_in_row = 0
        for part in parts:
            if current_x + part['Length'] + gap > plywood_size[0]:
                current_x, current_y = gap, current_y + max_y_in_row + gap
                max_y_in_row = 0
            if current_y + part['Height'] + gap > plywood_size[1]:
                sheet_layouts.append(current_sheet)
                current_sheet = {'parts': [], 'positions': []}
                current_y = gap
            current_sheet['parts'].append(part)
            current_sheet['positions'].append((current_x, current_y))
            current_x += part['Length'] + gap
            max_y_in_row = max(max_y_in_row, part['Height'])
        if current_sheet['parts']:
            sheet_layouts.append(current_sheet)
        return sheet_layouts

    def create_svg(self, sheet_layouts, project_folder, gap, material):
        for i, sheet in enumerate(sheet_layouts, start=1):
            # Filename now includes the material name
            output_path = os.path.join(project_folder, f"{material}_cut_{i}.svg")
            dwg = Drawing(output_path, size=(f'{self.stock_width.get()}mm', f'{self.stock_length.get()}mm'))
            for part, (x, y) in zip(sheet['parts'], sheet['positions']):
                dwg.add(dwg.rect(insert=(f'{x}mm', f'{y}mm'), size=(f"{part['Length']}mm", f"{part['Height']}mm"), stroke='black', fill='none'))
                text_x = x + part['Length'] / 2
                text_y = y + part['Height'] / 2
                dwg.add(dwg.text(part['Part Label'], insert=(f'{text_x}mm', f'{text_y}mm'), fill='black', text_anchor="middle", dominant_baseline="central", font_size='15px', font_family="Arial"))
                dimensions_text = f"{part['Length']}x{part['Height']}mm"
                dwg.add(dwg.text(dimensions_text, insert=(f'{text_x}mm', f'{text_y + 10}mm'), fill='black', text_anchor="middle", dominant_baseline="hanging", font_size='15px', font_family="Arial"))
            dwg.save()

    def display_instructions(self):
        instructions = """
        Instructions:
        
        The Excel file should have the following format
        
            Column A  | Column B | Column C | Column D  | Column E
           Part Label |  Length  |  Width   | Pcs Count | Material
            FrDoor    |   1800   |   450    |    2      | Hard Ply

        UNITS ARE ASSUMED MM. Do not include units in length or width.

        If you leave Quantity or Piece Count blank, it will assume 1.
        
        The "Gap Between Parts" is just that - will 'optimize' the
        layout of peices so there is at least that much space for your
        CNC router bit.  Round up to the 1/10th of a MM.

        Support my bad decisions by sending me a beer in venmo
        @Denver-Lancaster
        """
        instructions_window = tk.Toplevel(self.window)
        instructions_window.title("Instructions")
        instructions_text = scrolledtext.ScrolledText(instructions_window, width=100, height=50)
        instructions_text.insert(tk.INSERT, instructions)
        instructions_text.config(state=tk.DISABLED)
        instructions_text.pack(padx=10, pady=10)

    def open_output_directory(self):
        # You can retrieve the directory path from the output_entry widget
        directory_path = self.output_entry.get()

        # Alternatively, you can call the determine_project_folder method to get the latest directory path
        # directory_path = self.determine_project_folder()

        # Check if the directory path is not empty
        if directory_path:
            # Use subprocess to open the directory
            try:
                if os.name == 'nt':  # for Windows
                    os.startfile(directory_path)
                elif os.name == 'posix':  # for macOS and Linux
                    subprocess.Popen(['open', directory_path])
                else:  # fallback for other OS
                    subprocess.Popen(['xdg-open', directory_path])
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open directory: {e}")
        else:
            messagebox.showwarning("Warning", "No output directory selected or generated yet.")

    def save_output_folder_path(self, path):
        settings_path = os.path.join(os.path.dirname(__file__), 'last_output_folder.txt')
        try:
            with open(settings_path, 'w') as file:
                file.write(path)
        except Exception as e:
            print(f"Error saving last output folder path: {e}")

    def load_last_output_folder_path(self):
        settings_path = os.path.join(os.path.dirname(__file__), 'last_output_folder.txt')
        try:
            with open(settings_path, 'r') as file:
                return file.read().strip()
        except Exception:
            return None

if __name__ == "__main__":
    setup_resources()
    CutlistOptimizerGUI()
